# -*- coding: utf-8 -*-
import openpyxl
import json

# Открываем Excel файл
wb = openpyxl.load_workbook('РАСПИСАНИЕ 2025-2026.xlsx')

# Берем лист 7А
ws = wb['7А']

print("Первые 30 строк листа 7А:")
print("=" * 100)

for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, max_col=10), start=1):
    cells = []
    for cell in row:
        value = cell.value
        if value is not None:
            cells.append(str(value)[:50])  # Ограничиваем длину
        else:
            cells.append("")

    print(f"Строка {row_idx:2d}: {cells[:6]}")  # Первые 6 колонок

print("\n" + "=" * 100)
print("Сохраняю в файл для анализа...")

# Сохраняем полные данные в JSON
data = []
for row in ws.iter_rows(min_row=1, max_row=30, max_col=10):
    row_data = []
    for cell in row:
        row_data.append(str(cell.value) if cell.value is not None else "")
    data.append(row_data)

with open('excel_structure.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("Готово! Данные сохранены в excel_structure.json")
