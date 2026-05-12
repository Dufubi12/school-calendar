# -*- coding: utf-8 -*-
import openpyxl
import json
from datetime import datetime, timedelta

def parse_class_schedule(ws):
    """Парсит расписание класса из листа Excel"""
    days_schedule = {
        'Понедельник': [],
        'Вторник': [],
        'Среда': [],
        'Четверг': [],
        'Пятница': []
    }

    # Находим строку с днями недели
    day_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row and 'Понедельник' in str(row):
            day_row = row_idx
            break

    if not day_row:
        return days_schedule

    # Получаем дни недели и их позиции
    header_row = list(ws.iter_rows(min_row=day_row, max_row=day_row, values_only=True))[0]
    day_columns = {}

    for col_idx, cell_value in enumerate(header_row):
        if cell_value in days_schedule.keys():
            day_columns[col_idx] = cell_value

    # Парсим уроки
    for row in ws.iter_rows(min_row=day_row+1, max_row=day_row+15, values_only=True):
        if not row or not row[0]:
            continue

        time_str = str(row[0]).strip()

        # Проверяем что это время урока
        if ':' in time_str and '-' in time_str and not time_str.startswith('1'):
            continue
        if not any(char.isdigit() for char in time_str):
            continue
        if 'обед' in time_str.lower() or 'кружк' in time_str.lower():
            continue

        if ':' in time_str and '-' in time_str:
            # Это урок
            for col_idx, day_name in day_columns.items():
                if col_idx < len(row):
                    subject = row[col_idx]
                    if subject and str(subject).strip() and str(subject) != 'None':
                        subject_clean = str(subject).strip()
                        # Извлекаем предмет и учителя
                        if '   ' in subject_clean:
                            parts = subject_clean.split('   ')
                            subject_name = parts[0].strip()
                            teacher = parts[1].strip() if len(parts) > 1 else ''
                        else:
                            subject_name = subject_clean
                            teacher = ''

                        days_schedule[day_name].append({
                            'time': time_str,
                            'subject': subject_name,
                            'teacher': teacher
                        })

    return days_schedule

# Открываем Excel файл
wb = openpyxl.load_workbook('РАСПИСАНИЕ 2025-2026.xlsx')

# Классы для обработки
classes_to_parse = {
    '7А': '7А',
    '7Б': '7Б',
    '8А': '8А',
    '9А': '9А'
}

all_schedules = {}

for class_code, class_name in classes_to_parse.items():
    print(f'Обработка класса {class_name}...')
    ws = wb[class_code]
    schedule = parse_class_schedule(ws)
    all_schedules[class_name] = schedule

# Сохраняем в JSON
with open('parsed_schedules.json', 'w', encoding='utf-8') as f:
    json.dump(all_schedules, f, ensure_ascii=False, indent=2)

print('\nГотово! Расписание сохранено в parsed_schedules.json')
print(f'Обработано классов: {len(all_schedules)}')

# Показываем пример
print('\nПример данных для 7А класса (Понедельник):')
if '7А' in all_schedules and 'Понедельник' in all_schedules['7А']:
    for lesson in all_schedules['7А']['Понедельник'][:5]:
        print(f"  {lesson['time']}: {lesson['subject']} - {lesson['teacher']}")
