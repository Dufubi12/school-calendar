# -*- coding: utf-8 -*-
import openpyxl
import json

# Load workbook
wb = openpyxl.load_workbook('РАСПИСАНИЕ 2025-2026.xlsx')

# Save sheet names
with open('sheet_names.json', 'w', encoding='utf-8') as f:
    json.dump(wb.sheetnames, f, ensure_ascii=False, indent=2)

print('Sheet names saved to sheet_names.json')
print(f'Total sheets: {len(wb.sheetnames)}')
